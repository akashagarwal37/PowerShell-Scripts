import openpyxl
path = "sample1.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active
 
# Cell object is created by using sheet object's cell() method.
cell_obj = sheet_obj['A1': 'B6']
 
# Print value of cell object using the value attribute
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)
from openpyxl import Workbook
# Call a Workbook() method to create new spreadsheet
workbook = Workbook()
workbook.save(filename="sample.xlsx")
