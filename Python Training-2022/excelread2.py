import openpyxl
path = "sample.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active
 
# Getting the value of maximum rows and column
row = sheet_obj.max_row
column = sheet_obj.max_column
 
print("Total Rows:", row)
print("Total Columns:", column)
 
print("\nValue of first column")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)
print("\nValue of first row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")
